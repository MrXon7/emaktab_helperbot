import os
import io
import time
import asyncio
import logging
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, UploadFile, File, HTTPException, Depends
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from aiogram import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import settings
from bot import bot, dp
from database import init_db, get_db, User, Student
from auth import get_current_user
from emaktab_service import EmaktabService
from excel_parser import ExcelParser

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SEVEN_DAYS_MS = 7 * 24 * 60 * 60 * 1000

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: Supabase jadvallarini tekshirish/yaratish
    try:
        init_db()
    except Exception as e:
        logger.error(f"DB Startup error: {e}")

    # Webhook sozlash (agar token bo'lsa)
    if bot and settings.BOT_TOKEN:
        webhook_url = f"{settings.WEBAPP_URL.rstrip('/')}/webhook"
        try:
            await bot.set_webhook(url=webhook_url, drop_pending_updates=True)
            logger.info(f"Telegram Webhook o'rnatildi: {webhook_url}")
        except Exception as e:
            logger.warning(f"Webhook o'rnatilmadi: {e}")
    
    yield
    
    # Shutdown
    if bot and settings.BOT_TOKEN:
        try:
            await bot.delete_webhook()
            await bot.session.close()
            logger.info("Telegram Bot sessiyasi yopildi.")
        except Exception as e:
            logger.error(f"Botni to'xtatishda xatolik: {e}")

app = FastAPI(title="eMaktab Helper Multi-User", lifespan=lifespan)

# Statik fayllar va shablonlar
static_dir = os.path.join(BASE_DIR, "static")
templates_dir = os.path.join(BASE_DIR, "templates")

app.mount("/static", StaticFiles(directory=static_dir), name="static")
templates = Jinja2Templates(directory=templates_dir)

# Schemas
class StudentCreateOrUpdateRequest(BaseModel):
    name: str
    schoolName: str = "Maktab"
    grade: str = "1-A"
    login: str
    password: str

class StudentLoginRequest(BaseModel):
    id: str
    name: str
    login: str
    password: str
    schoolName: str = "Maktab"
    grade: str = "1-A"

@app.get("/", response_class=HTMLResponse)
async def serve_webapp(request: Request):
    """Telegram Mini App bosh sahifasi"""
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"request": request}
    )

@app.get("/health")
async def health_check():
    return {"status": "ok", "app": "eMaktab Helper Multi-User", "database": "Supabase PostgreSQL"}

@app.get("/api/me")
async def get_me(user: User = Depends(get_current_user)):
    """Foydalanuvchi ma'lumotlari"""
    return {
        "id": user.id,
        "telegram_id": user.telegram_id,
        "name": user.first_name,
        "username": user.username
    }

@app.get("/api/download-template")
async def download_template():
    """Namuna Excel (.xlsx) faylini yuklab olish"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"

    headers = ["F.I.Sh (Ism Familiya)", "Maktab", "Sinf", "Login", "Parol"]
    ws.append(headers)

    samples = [
        ["Aliyev Vali G'aniyevich", "56-Maktab", "5-A", "ali_valiyev_5a", "Parol123!"],
        ["Karimova Madina Rustam qizi", "56-Maktab", "5-A", "madina_k_5a", "Madina2026"],
        ["Toshmatov Dilshod Akrom o'g'li", "56-Maktab", "6-B", "dilshod_t_6b", "Dilshod_123"]
    ]
    for row in samples:
        ws.append(row)

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 28

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=emaktab_oquvchilar_namuna.xlsx"}
    )

@app.get("/api/students")
async def get_students(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Joriy foydalanuvchining o'quvchilari (7 kunlik muddati tekshirilgan holda)"""
    students = db.query(Student).filter(Student.user_id == user.id).order_by(Student.created_at.desc()).all()
    
    now_ms = int(time.time() * 1000)
    has_expired = False

    result = []
    for s in students:
        # 7 kunlik tekshiruv
        if s.status == "success" and s.success_at:
            if (now_ms - s.success_at) >= SEVEN_DAYS_MS:
                s.status = "pending"
                s.message = "1 hafta o'tgani sababli qayta kutilmoqda"
                has_expired = True
        result.append(s.to_dict())

    if has_expired:
        db.commit()

    return {"students": result}

@app.post("/api/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Excel yuklash va foydalanuvchi nomiga Supabase'ga saqlash"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Faqat .xlsx yoki .xls fayllar qabul qilinadi")
    
    try:
        content = await file.read()
        parsed_students = ExcelParser.parse_excel_bytes(content)

        saved_students = []
        for item in parsed_students:
            student = Student(
                id=item["id"],
                user_id=user.id,
                name=item["name"],
                school_name=item["schoolName"],
                grade=item["grade"],
                login=item["login"],
                password=item["password"],
                status="pending",
                message=""
            )
            db.add(student)
            saved_students.append(student.to_dict())

        db.commit()
        return {"success": True, "count": len(saved_students), "students": saved_students}
    except Exception as e:
        db.rollback()
        logger.exception("Excel parsing va DB saqlash xatosi")
        raise HTTPException(status_code=500, detail=f"Faylni saqlashda xatolik: {str(e)}")

@app.post("/api/students")
async def create_student(
    req: StudentCreateOrUpdateRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Qo'lda yangi o'quvchi qo'shish"""
    student_id = f"std_{int(time.time() * 1000)}_{os.urandom(2).hex()}"
    student = Student(
        id=student_id,
        user_id=user.id,
        name=req.name.strip(),
        school_name=req.schoolName.strip() or "Maktab",
        grade=req.grade.strip() or "1-A",
        login=req.login.strip(),
        password=req.password.strip(),
        status="pending",
        message=""
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return {"success": True, "student": student.to_dict()}

@app.put("/api/students/{student_id}")
async def update_student(
    student_id: str,
    req: StudentCreateOrUpdateRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """O'quvchini tahrirlash (faqat o'ziga tegishlisini)"""
    student = db.query(Student).filter(Student.id == student_id, Student.user_id == user.id).first()
    if not student:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")

    student.name = req.name.strip()
    student.school_name = req.schoolName.strip() or "Maktab"
    student.grade = req.grade.strip() or "1-A"
    student.login = req.login.strip()
    student.password = req.password.strip()
    student.status = "pending"
    student.message = "Ma'lumotlar tahrirlandi"
    
    db.commit()
    db.refresh(student)
    return {"success": True, "student": student.to_dict()}

@app.delete("/api/students/{student_id}")
async def delete_student(
    student_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """O'quvchini o'chirish"""
    student = db.query(Student).filter(Student.id == student_id, Student.user_id == user.id).first()
    if not student:
        raise HTTPException(status_code=404, detail="O'quvchi topilmadi")

    db.delete(student)
    db.commit()
    return {"success": True, "message": "O'quvchi o'chirildi"}

@app.post("/api/login-single")
async def login_single(
    student_req: StudentLoginRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bitta o'quvchiga emaktab.uz orqali kirish va natijani Supabase bazasida yangilash"""
    result = await EmaktabService.process_student_login(student_req.model_dump())
    
    # Bazadagi statusni yangilash
    student = db.query(Student).filter(Student.id == student_req.id, Student.user_id == user.id).first()
    if student:
        student.status = result["status"]
        student.message = result.get("message", "")
        if result["status"] == "success":
            student.success_at = int(time.time() * 1000)
        db.commit()

    return result

@app.post("/webhook")
async def telegram_webhook(request: Request):
    """Telegram Bot Webhook endpointi"""
    if not bot:
        return JSONResponse({"status": "Bot token not configured"}, status_code=200)
    
    try:
        data = await request.json()
        update = types.Update(**data)
        await dp.feed_update(bot, update)
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"Webhook xatosi: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host=settings.HOST, port=settings.PORT, reload=True)
